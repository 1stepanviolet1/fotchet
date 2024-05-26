from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


class OtchetTable(Workbook):
    A1_L1 = '''Отчет о результатах самостоятельной работы обучающегося по дисциплинам
"Физическая культура" или "Элективные курсы по физической культуре и спорту"'''

    A3 = '''Студ. билет'''

    B3_E3 = '''ФИО'''

    F3 = '''Группа'''

    G3_I3 = '''Спортивное отделение'''

    J3_L3 = '''Преподаватель'''

    A7_F7 = '''Шахматная партия'''
    G7_L7 = A7_F7

    A8_F8 = '''Играна во "%s" туре %s'''
    G8_L8 = A8_F8

    A9_F9 = '''Белые: %s'''
    G9_L9 = A9_F9

    A10_F10 = '''Чёрные: %s'''
    G10_L10 = A10_F10

    pttrn_battle_intro = '''№ Белые Чёрные'''

    pttrn_res = '''Результат:'''

    def __init__(self, iso_dates=False) -> None:
        super().__init__(iso_dates=iso_dates)

        _sheet = self.active

        _sheet.row_dimensions[1].height = 55
        _sheet.row_dimensions[3].height = 25
        
        for col in 'ADGJ':
            _sheet.column_dimensions[col].width = 12
    
    def init_A1_L1(self):
        _sheet = self.active

        _sheet.merge_cells("A1:L1")

        _sheet["A1"].font = Font(size=14, bold=True)
        _sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        _sheet["A1"].value = self.A1_L1
    
    def init_A3(self):
        _sheet = self.active

        _sheet["A3"].font = Font(bold=True)
        _sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")
        _sheet["A3"].value = self.A3

        _sheet["A4"].alignment = Alignment(horizontal="center", vertical="center")
    
    def init_B3_E3(self):
        _sheet = self.active

        _sheet.merge_cells("B3:E3")
        _sheet.merge_cells("B4:E4")

        _sheet["B3"].font = Font(bold=True)
        _sheet["B3"].alignment = Alignment(horizontal="center", vertical="center")
        _sheet["B3"].value = self.B3_E3

        _sheet["B4"].alignment = Alignment(horizontal="center", vertical="center")
    
    def init_F3(self):
        _sheet = self.active

        _sheet["F3"].font = Font(bold=True)
        _sheet["F3"].alignment = Alignment(horizontal="center", vertical="center")
        _sheet["F3"].value = self.F3

        _sheet["F4"].alignment = Alignment(horizontal="center", vertical="center")

    def init_G3_I3(self):
        _sheet = self.active

        _sheet.merge_cells("G3:I3")
        _sheet.merge_cells("G4:I4")
        
        _sheet["G3"].font = Font(bold=True)
        _sheet["G3"].alignment = Alignment(horizontal="center", vertical="center")
        _sheet["G3"].value = self.G3_I3

        _sheet["G4"].alignment = Alignment(horizontal="center", vertical="center")
        _sheet["G4"].value = "Шахматы"
    
    def init_J3_L3(self):
        _sheet = self.active

        _sheet.merge_cells("J3:L3")
        _sheet.merge_cells("J4:L4")
        
        _sheet["J3"].font = Font(bold=True)
        _sheet["J3"].alignment = Alignment(horizontal="center", vertical="center")
        _sheet["J3"].value = self.J3_L3

        _sheet["J4"].alignment = Alignment(horizontal="center", vertical="center")
        _sheet["J4"].value = "Иванов С.В."
    
    def set_borders(self):
        _sheet = self.active

        for col in "ABCDEFGHIJK":
            _sheet[f'{col}3'].border = Border(
                left=Side(style='dashed'),
                right=Side(style='dashed'),
                top=Side(style='medium'),
                bottom=Side(style='dashed')
            )

            _sheet[f'{col}4'].border = Border(
                left=Side(style='dashed'),
                right=Side(style='dashed'),
                top=Side(style='dashed'),
                bottom=Side(style='medium')
            )
        
        _sheet['L3'].border = Border(
            left=Side(style='dashed'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='dashed')
        )

        _sheet['L4'].border = Border(
            left=Side(style='dashed'),
            right=Side(style='medium'),
            top=Side(style='dashed'),
            bottom=Side(style='medium')
        )
    
    def set_student_id(self, _val: str):
        self.active["A4"].value = int(_val)

    def set_fio(self, _val: str):
        self.active["B4"].value = _val
    
    def set_group(self, _val: str):
        self.active["F4"].value = int(_val)

    def make_intro(self, *, student_id: str, fio: str):
        self.init_A1_L1()
        self.init_A3()
        self.init_B3_E3()
        self.init_F3()
        self.init_G3_I3()
        self.init_J3_L3()

        self.set_student_id(student_id)
        self.set_fio(fio)
        self.set_group(student_id[:4])

        self.set_borders()

    def init_A7_F7(self):
        _sheet = self.active

        _sheet.merge_cells("A7:F7")

        _sheet["A7"].font = Font(bold=True)
        _sheet["A7"].alignment = Alignment(horizontal="center", vertical="center")
        _sheet["A7"].value = self.A7_F7
    
    def init_A8_F8(self, _tour: str, _date: str):
        _sheet = self.active

        _sheet.merge_cells("A8:F8")

        _sheet["A8"].alignment = Alignment(horizontal="center", vertical="center")
        _sheet["A8"].value = self.A8_F8 % (_tour, _date)
    
    def init_A9_F9(self, _your_name: str):
        _sheet = self.active

        _sheet.merge_cells("A9:F9")

        _sheet["A9"].alignment = Alignment(horizontal="left", vertical="center")
        _sheet["A9"].value = self.A9_F9 % _your_name
    
    def init_A10_F10(self, _opponent_name: str):
        _sheet = self.active

        _sheet.merge_cells("A10:F10")

        _sheet["A10"].alignment = Alignment(horizontal="left", vertical="center")
        _sheet["A10"].value = self.A10_F10 % _opponent_name
    
    def make_battle(self, 
                      data: dict[list[str]], 
                      *, 
                      offset: int):
        x_offset = offset
        y_offset = 0
        _white_won: bool

        len_battle = len(data.get('white'))

        if len_battle > 72:
            print("Error: простите, но в партии больше 72 ходов")
            exit(1)
        
        if len_battle > len(data.get('black')):
            _white_won = True
        else:
            _white_won = False

        self.create_cols_of_moves(offset=x_offset)
        self.create_battle_numbers(x_offset=x_offset)
        self.create_battle_result(_white_won, offset=x_offset)
        
        for i in range(len_battle):
            if i == 36:
                x_offset += 3
                y_offset -= 36
                self.create_cols_of_moves(offset=x_offset)
                self.create_battle_numbers(offset=36, x_offset=x_offset)

            w_step = data['white'][i]
            try:
                b_step = data['black'][i]
            except IndexError:
                b_step = ''

            self.create_battle_move((w_step, b_step), x_offset=x_offset, y_offset=i+y_offset)
    
    def create_cols_of_moves(self, *, offset: int): # start: A11
        _sheet = self.active

        for i in 1, 2, 3:
            _cell = _sheet.cell(row=11, column=i+offset)
            _cell.alignment = Alignment(horizontal="center", vertical="center")
            _cell.value = self.pttrn_battle_intro.split()[i-1]
    
    def create_battle_result(self, _white_won, *, offset: int): # start: D48
        if offset < 0:
            raise ValueError(f"offset must be positive, not {offset}")
    
        _sheet = self.active

        _cell = _sheet.cell(row=48, column=4+offset)
        _cell.alignment = Alignment(horizontal="center", vertical="center")
        _cell.value = self.pttrn_res

        _cell = _sheet.cell(row=48, column=4+offset+1)
        _cell.alignment = Alignment(horizontal="center", vertical="center")
        _cell.value = int(_white_won)

        _cell = _sheet.cell(row=48, column=4+offset+2)
        _cell.alignment = Alignment(horizontal="center", vertical="center")
        _cell.value = int(not _white_won)
    
    def create_battle_numbers(self, *, offset=0, x_offset): # start: A12
        _sheet = self.active

        for i in range(36):
            _cell = _sheet.cell(row=12+i, column=1+x_offset)
            _cell.alignment = Alignment(horizontal="center", vertical="center")
            _cell.font = Font(bold=True)
            _cell.value = i + 1 + offset
    
    def create_battle_move(self, moves, *, x_offset: int, y_offset: int): # start: B12
        for i in 0, 1:
            self.set_move(moves[i], x=2+x_offset+i, y=12+y_offset)
    
    def set_move(self, move: str, *, x, y):
        _sheet = self.active
        _cell = _sheet.cell(row=y, column=x)
        _cell.alignment = Alignment(horizontal="left", vertical="center")
        _cell.value = move

    def make_1_battle(self, 
                      data: dict[list[str]], 
                      *, 
                      tour: str, 
                      date: str, 
                      your_name: str, 
                      opponent_name: str):
        self.init_A7_F7()
        self.init_A8_F8(tour, date)
        self.init_A9_F9(your_name)
        self.init_A10_F10(opponent_name)

        self.make_battle(data, offset=0)

    def init_G7_L7(self):
        _sheet = self.active

        _sheet.merge_cells("G7:L7")

        _sheet["G7"].font = Font(bold=True)
        _sheet["G7"].alignment = Alignment(horizontal="center", vertical="center")
        _sheet["G7"].value = self.G7_L7
    
    def init_G8_L8(self, _tour: str, _date: str):
        _sheet = self.active

        _sheet.merge_cells("G8:L8")

        _sheet["G8"].alignment = Alignment(horizontal="center", vertical="center")
        _sheet["G8"].value = self.G8_L8 % (_tour, _date)
    
    def init_G9_L9(self, _player1: str):
        _sheet = self.active

        _sheet.merge_cells("G9:L9")

        _sheet["G9"].alignment = Alignment(horizontal="left", vertical="center")
        _sheet["G9"].value = self.G9_L9 % _player1
    
    def init_G10_L10(self, _player2: str):
        _sheet = self.active

        _sheet.merge_cells("G10:L10")

        _sheet["G10"].alignment = Alignment(horizontal="left", vertical="center")
        _sheet["G10"].value = self.G10_L10 % _player2
    
    def make_2_battle(self, 
                      data: dict[list[str]], 
                      *, 
                      tour: str, 
                      date: str, 
                      your_name: str, 
                      opponent_name: str):
        self.init_G7_L7()
        self.init_G8_L8(tour, date)
        self.init_G9_L9(opponent_name)
        self.init_G10_L10(your_name)

        self.make_battle(data, offset=6) 
